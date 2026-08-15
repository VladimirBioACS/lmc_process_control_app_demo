"""Report builder module for LMC process control application."""

from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from pathlib import Path
import random
from typing import Any

from openpyxl import Workbook, load_workbook


TITLE_SHEET = "Звіт_титульний_аркуш"
WORK_ZONE_SHEET = "Звіт_Технологічні_умови_в_робочій_зоні_установки"
GRADIENT_SHEET = "Звіт_температурний_градієнт"
WORK_ZONE_SHEET_FALLBACK = "Звіт_Технологічні_умови_в_робоч"

DEFAULT_TEMPLATE_PATH = (
    Path(__file__).resolve().parent / "report_templates" / "uvnk_8p_report_template.xlsx"
)


@dataclass(slots=True)
class ReportStubData:
    """Container with stub values used to fill all report sheets."""

    operator_name: str
    lead_engineer_name: str
    smelting_number: int
    report_date: date
    furnace_unit_number: int
    process_start_time: time
    process_end_time: time
    part_name: str
    part_number: int

    lmc_start_sec: int
    lmc_end_sec: int
    vacuum_start: int
    vacuum_end: int
    mold_initial_temperature_c: int
    chamber_rise_speed_c_per_min: int
    direct_crystallization_speed: int
    pour_area: int
    crystallization_begin_area: int
    border_mold_lowering_area: int
    metal_pouring_temperature_c: int
    cold_junction_temperature_c: int
    ambient_temperature_c: int

    ideal_gradient: int
    min_gradient: int
    max_gradient: int
    final_gradient: int
    thermal_couple_count: int
    thermal_couple_distances: list[int]

    lmc_time_seconds: list[int]
    furnace_heater_temperatures_c: list[int]
    aluminium_coolant_temperatures_c: list[int]
    aluminium_coolant_heater_temperatures_c: list[int]

    gradient_values: list[float]
    cooling_rates: list[int]
    linear_fit_quality: list[float]

    tc_channel_1_temperatures_c: list[float]
    tc_channel_2_temperatures_c: list[float]
    tc_channel_3_temperatures_c: list[float]
    tc_channel_4_temperatures_c: list[float]
    tc_channel_5_temperatures_c: list[float]


def _linear_series(start: float, end: float, count: int, precision: int = 3) -> list[float]:
    """Build a linear inclusive series from start to end."""

    if count <= 1:
        return [round(start, precision)]

    step = (end - start) / (count - 1)
    return [round(start + step * i, precision) for i in range(count)]


def generate_stub_data(
    *,
    start_sec: int = 1800,
    end_sec: int = 3500,
    random_seed: int = 42,
) -> ReportStubData:
    """Generate deterministic stub data for report generation and verification."""

    if end_sec < start_sec:
        raise ValueError("end_sec must be greater than or equal to start_sec")

    rng = random.Random(random_seed)
    lmc_time_seconds = list(range(start_sec, end_sec + 1))
    points_count = len(lmc_time_seconds)

    return ReportStubData(
        operator_name="Operator Name",
        lead_engineer_name="Operator Name",
        smelting_number=1,
        report_date=date.today(),
        furnace_unit_number=1,
        process_start_time=time(hour=0, minute=0, second=0),
        process_end_time=time(hour=2, minute=20, second=20),
        part_name="лопатка ГТД",
        part_number=1,
        lmc_start_sec=start_sec,
        lmc_end_sec=end_sec,
        vacuum_start=8,
        vacuum_end=8,
        mold_initial_temperature_c=678,
        chamber_rise_speed_c_per_min=15,
        direct_crystallization_speed=10,
        pour_area=50,
        crystallization_begin_area=125,
        border_mold_lowering_area=250,
        metal_pouring_temperature_c=1623,
        cold_junction_temperature_c=20,
        ambient_temperature_c=22,
        ideal_gradient=20,
        min_gradient=19,
        max_gradient=20,
        final_gradient=20,
        thermal_couple_count=5,
        thermal_couple_distances=[20, 20, 20, 20, 20],
        lmc_time_seconds=lmc_time_seconds,
        furnace_heater_temperatures_c=[rng.randint(1548, 1557) for _ in lmc_time_seconds],
        aluminium_coolant_temperatures_c=[rng.randint(750, 755) for _ in lmc_time_seconds],
        aluminium_coolant_heater_temperatures_c=[rng.randint(750, 755) for _ in lmc_time_seconds],
        gradient_values=_linear_series(1.0, 20.0, points_count, precision=3),
        cooling_rates=[rng.randint(5, 20) for _ in lmc_time_seconds],
        linear_fit_quality=[round(rng.uniform(0.01, 1.0), 3) for _ in lmc_time_seconds],
        tc_channel_1_temperatures_c=_linear_series(1375.0, 550.0, points_count, precision=2),
        tc_channel_2_temperatures_c=_linear_series(1360.0, 600.0, points_count, precision=2),
        tc_channel_3_temperatures_c=_linear_series(1360.0, 650.0, points_count, precision=2),
        tc_channel_4_temperatures_c=_linear_series(1360.0, 740.0, points_count, precision=2),
        tc_channel_5_temperatures_c=_linear_series(1360.0, 7960.0, points_count, precision=2),
    )


def _fill_column(
    worksheet: Any,
    *,
    start_row: int,
    column_letter: str,
    values: list[Any],
) -> None:
    """Write values down a single worksheet column starting from start_row."""

    for index, value in enumerate(values):
        worksheet[f"{column_letter}{start_row + index}"] = value


def _resolve_sheet_name(workbook: Workbook, preferred_name: str, fallback_name: str | None = None) -> str:
    """Return matching worksheet name from template, supporting known fallback names."""

    if preferred_name in workbook.sheetnames:
        return preferred_name
    if fallback_name and fallback_name in workbook.sheetnames:
        return fallback_name
    raise KeyError(
        f"Template is missing required sheet '{preferred_name}'"
        + (f" (or fallback '{fallback_name}')" if fallback_name else "")
    )


def _fill_title_sheet(workbook: Workbook, data: ReportStubData) -> None:
    sheet = workbook[_resolve_sheet_name(workbook, TITLE_SHEET)]

    sheet["B4"] = data.operator_name
    sheet["B5"] = data.lead_engineer_name
    sheet["B7"] = data.smelting_number
    sheet["B8"] = data.report_date
    sheet["B8"].number_format = "DD/MM/YY"
    sheet["B9"] = data.furnace_unit_number
    sheet["B10"] = data.process_start_time
    sheet["B10"].number_format = "hh:mm:ss"
    sheet["B11"] = data.process_end_time
    sheet["B11"].number_format = "hh:mm:ss"
    sheet["B12"] = "=B11-B10"
    sheet["B12"].number_format = "[h]:mm:ss"
    sheet["B14"] = data.part_name
    sheet["B15"] = data.part_number
    sheet["B16"] = data.furnace_unit_number


def _fill_work_zone_sheet(workbook: Workbook, data: ReportStubData) -> None:
    sheet = workbook[_resolve_sheet_name(workbook, WORK_ZONE_SHEET, WORK_ZONE_SHEET_FALLBACK)]

    sheet["B5"] = data.lmc_start_sec
    sheet["B6"] = data.lmc_end_sec
    sheet["B8"] = data.vacuum_start
    sheet["B9"] = data.vacuum_end
    sheet["B11"] = data.mold_initial_temperature_c
    sheet["B12"] = data.chamber_rise_speed_c_per_min
    sheet["B14"] = data.direct_crystallization_speed
    sheet["B16"] = data.pour_area
    sheet["B17"] = data.crystallization_begin_area
    sheet["B18"] = data.border_mold_lowering_area
    sheet["B20"] = data.metal_pouring_temperature_c
    sheet["B22"] = data.cold_junction_temperature_c
    sheet["B23"] = data.ambient_temperature_c

    start_row = 6
    _fill_column(sheet, start_row=start_row, column_letter="E", values=data.lmc_time_seconds)
    _fill_column(
        sheet,
        start_row=start_row,
        column_letter="F",
        values=data.furnace_heater_temperatures_c,
    )
    _fill_column(
        sheet,
        start_row=start_row,
        column_letter="H",
        values=data.aluminium_coolant_temperatures_c,
    )
    _fill_column(
        sheet,
        start_row=start_row,
        column_letter="I",
        values=data.aluminium_coolant_heater_temperatures_c,
    )


def _fill_gradient_sheet(workbook: Workbook, data: ReportStubData) -> None:
    sheet = workbook[_resolve_sheet_name(workbook, GRADIENT_SHEET)]

    sheet["B5"] = data.ideal_gradient
    sheet["B6"] = data.min_gradient
    sheet["B7"] = data.max_gradient
    sheet["B9"] = data.final_gradient
    sheet["B12"] = data.thermal_couple_count

    sheet["B14"] = data.thermal_couple_distances[0]
    sheet["B15"] = data.thermal_couple_distances[1]
    sheet["B16"] = data.thermal_couple_distances[2]
    sheet["B17"] = data.thermal_couple_distances[3]
    sheet["B18"] = data.thermal_couple_distances[4]

    start_row = 6
    _fill_column(sheet, start_row=start_row, column_letter="E", values=data.lmc_time_seconds)
    _fill_column(sheet, start_row=start_row, column_letter="F", values=data.gradient_values)
    _fill_column(sheet, start_row=start_row, column_letter="G", values=data.cooling_rates)
    _fill_column(sheet, start_row=start_row, column_letter="H", values=data.linear_fit_quality)

    _fill_column(sheet, start_row=start_row, column_letter="J", values=data.lmc_time_seconds)
    _fill_column(
        sheet,
        start_row=start_row,
        column_letter="K",
        values=data.tc_channel_1_temperatures_c,
    )
    _fill_column(
        sheet,
        start_row=start_row,
        column_letter="L",
        values=data.tc_channel_2_temperatures_c,
    )
    _fill_column(
        sheet,
        start_row=start_row,
        column_letter="M",
        values=data.tc_channel_3_temperatures_c,
    )
    _fill_column(
        sheet,
        start_row=start_row,
        column_letter="N",
        values=data.tc_channel_4_temperatures_c,
    )
    _fill_column(
        sheet,
        start_row=start_row,
        column_letter="O",
        values=data.tc_channel_5_temperatures_c,
    )


def fill_report_workbook(workbook: Workbook, data: ReportStubData) -> None:
    """Fill workbook sheets while preserving existing charts/formulas/layout."""

    _resolve_sheet_name(workbook, TITLE_SHEET)
    _resolve_sheet_name(workbook, WORK_ZONE_SHEET, WORK_ZONE_SHEET_FALLBACK)
    _resolve_sheet_name(workbook, GRADIENT_SHEET)

    _fill_title_sheet(workbook, data)
    _fill_work_zone_sheet(workbook, data)
    _fill_gradient_sheet(workbook, data)


def build_report(
    template_path: str | Path = DEFAULT_TEMPLATE_PATH,
    output_path: str | Path | None = None,
    stub_data: ReportStubData | None = None,
) -> Path:
    """Build report from template and return generated file path."""

    template = Path(template_path)
    if not template.exists():
        raise FileNotFoundError(f"Template file was not found: {template}")

    data = stub_data if stub_data is not None else generate_stub_data()
    workbook = load_workbook(filename=template)
    fill_report_workbook(workbook, data)

    if output_path is None:
        output_dir = template.parent / "generated"
        output_dir.mkdir(parents=True, exist_ok=True)
        output_file = output_dir / f"uvnk_8p_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    else:
        output_file = Path(output_path)
        output_file.parent.mkdir(parents=True, exist_ok=True)

    workbook.save(output_file)
    return output_file


if __name__ == "__main__":
    generated_path = build_report()
    print(f"Report generated: {generated_path}")